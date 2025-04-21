import re
import pandas as pd
import psycopg2
from psycopg2.extras import execute_batch, execute_values
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QVBoxLayout, 
    QWidget, QFileDialog, QMessageBox, QProgressBar, QLabel,
    QTableWidget, QTableWidgetItem, QHeaderView
)
from PySide6.QtCore import QThread, Signal, Qt, QTimer
import chardet
import os
import glob
import traceback
import time
import logging
from typing import List, Dict, Optional, Set, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import wraps
import csv
import hashlib
from dataclasses import dataclass
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ==================== КОНФИГУРАЦИЯ ====================
CONFIG = {
    'database': {
        'host': 'localhost',
        'database': 'server_logs',
        'user': 'postgres',
        'password': '123',
        'port': '5432'
    },
    'batch_size': 100000,
    'log_sample_size': 5,
    'max_workers': max(4, os.cpu_count() - 2),
    'max_table_rows': 1000,
    'hash_salt': 'log_analyzer_salt_2023',
    'max_batch_rows': 2000000,
    'prefetch_ips': 100000,
    'connection_pool_size': 5
}

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='app.log',
    filemode='w'
)

def log_operation(message: str):
    logging.info(message)
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}")

# ==================== УТИЛИТЫ ====================
def timeit(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        elapsed = time.time() - start
        log_operation(f"{func.__name__} выполнена за {elapsed:.2f} сек")
        if args and isinstance(args[0], MainWindow):
            args[0].update_time_label(f"{func.__name__}: {elapsed:.2f} сек")
        return result
    return wrapper

def calculate_log_hash(log: dict, salt: str) -> str:
    hash_str = f"{log['ip']}_{log['datetime']}_{log['method']}_{log['api']}_{log['status']}"
    return hashlib.sha256((hash_str + salt).encode()).hexdigest()

# ==================== МОДЕЛИ ДАННЫХ ====================
@dataclass
class LogRecord:
    ip: str
    datetime: datetime
    method: str
    api: str
    status: int
    bytes: int
    hash: str

# ==================== ПАРСЕР ЛОГОВ ====================
class FastLogParser:
    LOG_PATTERN = re.compile(
        r'^(?P<ip>\S+)\s+\S+\s+\S+\s+\[(?P<datetime>\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}\s\+\d{4})\]\s+'
        r'"(?P<method>\w+)\s+(?P<api>\S+)\s+HTTP/\d\.\d"\s+'
        r'(?P<status>\d{3})\s+(?P<bytes>\d+)\s+'
        r'"(?P<referer>[^"]*)"\s+"(?P<user_agent>[^"]*)"\s+'
        r'(?P<response_time>\d+)'
    )
    
    DATE_FORMATS = [
        '%Y-%m-%d %H:%M:%S %z',
        '%d/%b/%Y:%H:%M:%S %z',
        '%d/%b/%Y:%H:%M:%S',
        '%d/%b/%Y %H:%M:%S %z',
        '%d-%b-%Y:%H:%M:%S %z',
        '%d/%m/%Y:%H:%M:%S %z',
        '%d-%m-%Y:%H:%M:%S %z'
    ]
    
    @classmethod
    def parse_datetime(cls, dt_str: str) -> Optional[datetime]:
        for fmt in cls.DATE_FORMATS:
            try:
                return datetime.strptime(dt_str, fmt)
            except ValueError:
                continue
        log_operation(f"Не удалось распарсить дату: {dt_str}")
        return None
    
    @classmethod
    def parse_line(cls, line: str, salt: str) -> Optional[LogRecord]:
        try:
            line = line.strip()
            if not line:
                return None
                
            match = cls.LOG_PATTERN.search(line)
            if not match:
                log_operation(f"Не совпало с шаблоном: {line[:200]}...")
                return None
            
            log_data = match.groupdict()
            
            dt = cls.parse_datetime(log_data['datetime'])
            if not dt:
                return None
            
            bytes_sent = int(log_data['bytes']) if log_data['bytes'].isdigit() else 0
            
            return LogRecord(
                ip=log_data['ip'],
                datetime=dt,
                method=log_data['method'],
                api=log_data['api'],
                status=int(log_data['status']),
                bytes=bytes_sent,
                hash=calculate_log_hash(log_data, salt)
            )
        except Exception as e:
            log_operation(f"Ошибка парсинга строки: {e}\nСтрока: {line[:200]}...")
            return None

    @classmethod
    def parse_file(cls, file_path: str, salt: str, seen_hashes: Set[str]) -> Tuple[List[dict], Set[str]]:
        try:
            with open(file_path, 'rb') as f:
                raw = f.read(10000)
                encoding = chardet.detect(raw)['encoding'] or 'utf-8'
            
            unique_logs = []
            new_hashes = set()
            parsed_count = 0
            error_count = 0
            
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                for i, line in enumerate(f, 1):
                    record = cls.parse_line(line, salt)
                    if record:
                        parsed_count += 1
                        if record.hash not in seen_hashes:
                            unique_logs.append({
                                'ip': record.ip,
                                'datetime': record.datetime,
                                'method': record.method,
                                'api': record.api,
                                'status': record.status,
                                'bytes': record.bytes,
                                'is_error': record.status >= 400,
                                'hash': record.hash
                            })
                            new_hashes.add(record.hash)
                    else:
                        error_count += 1
                        
                    if i % 100000 == 0:
                        log_operation(f"Обработано {i} строк, успешно {parsed_count}, ошибок {error_count}, новых {len(unique_logs)}")
            
            log_operation(
                f"Файл {file_path} обработан: "
                f"успешно {parsed_count}, ошибок {error_count}, "
                f"новых записей {len(unique_logs)}"
            )
            return unique_logs, new_hashes
        except Exception as e:
            log_operation(f"Ошибка парсинга файла {file_path}: {str(e)}")
            return [], set()

# ==================== БАЗА ДАННЫХ ====================
class DatabaseManager:
    _connection_pool = None

    @classmethod
    def _get_connection(cls):
        if cls._connection_pool is None:
            cls._connection_pool = []
            for _ in range(CONFIG['connection_pool_size']):
                conn = psycopg2.connect(**CONFIG['database'])
                cls._connection_pool.append(conn)
        return cls._connection_pool.pop(0)

    @classmethod
    def _return_connection(cls, conn):
        cls._connection_pool.append(conn)

    @staticmethod
    @timeit
    def clear_database():
        conn = None
        try:
            conn = psycopg2.connect(**CONFIG['database'])
            conn.autocommit = False
            cursor = conn.cursor()
            
            cursor.execute("DROP TABLE IF EXISTS requests CASCADE")
            cursor.execute("DROP TABLE IF EXISTS ip_addresses CASCADE")
            cursor.execute("DROP TABLE IF EXISTS log_hashes CASCADE")
            cursor.execute("DROP MATERIALIZED VIEW IF EXISTS daily_stats CASCADE")
            
            conn.commit()
            log_operation("База данных успешно очищена")
            return True
        except Exception as e:
            log_operation(f"Ошибка при очистке БД: {e}")
            if conn:
                conn.rollback()
            return False
        finally:
            if conn:
                conn.close()

    @staticmethod
    @timeit
    def create_schema():
        conn = None
        try:
            conn = psycopg2.connect(**CONFIG['database'])
            conn.autocommit = False
            cursor = conn.cursor()
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS log_hashes (
                    hash VARCHAR(64) PRIMARY KEY,
                    processed_at TIMESTAMP DEFAULT NOW()
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS ip_addresses (
                    id SERIAL PRIMARY KEY,
                    ip_address VARCHAR(15) UNIQUE NOT NULL,
                    total_requests INTEGER DEFAULT 0,
                    last_seen TIMESTAMP WITH TIME ZONE,
                    is_suspicious BOOLEAN DEFAULT FALSE
                );
                
                CREATE TABLE IF NOT EXISTS requests (
                    id SERIAL PRIMARY KEY,
                    ip_id INTEGER REFERENCES ip_addresses(id),
                    datetime TIMESTAMP WITH TIME ZONE NOT NULL,
                    request_method VARCHAR(10) NOT NULL,
                    api_path TEXT NOT NULL,
                    protocol VARCHAR(5) NOT NULL,
                    status_code INTEGER NOT NULL,
                    bytes_sent INTEGER NOT NULL,
                    referer TEXT,
                    user_agent TEXT,
                    response_time INTEGER,
                    is_error BOOLEAN DEFAULT FALSE,
                    is_attack BOOLEAN DEFAULT FALSE
                );
            """)
            
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_requests_datetime ON requests(datetime);
                CREATE INDEX IF NOT EXISTS idx_requests_ip_id ON requests(ip_id);
                CREATE INDEX IF NOT EXISTS idx_ip_addresses_ip ON ip_addresses(ip_address);
            """)
            
            cursor.execute("""
                CREATE MATERIALIZED VIEW IF NOT EXISTS daily_stats AS
                SELECT 
                    DATE(datetime) as day,
                    COUNT(*) as total_requests,
                    SUM(CASE WHEN is_error THEN 1 ELSE 0 END) as errors,
                    AVG(response_time) as avg_response_time
                FROM requests
                GROUP BY DATE(datetime)
                ORDER BY day;
                
                CREATE UNIQUE INDEX IF NOT EXISTS idx_daily_stats_day_unique ON daily_stats(day);
            """)
            
            conn.commit()
            log_operation("Схема БД успешно проверена/создана")
            return True
        except Exception as e:
            log_operation(f"Ошибка при создании схемы БД: {e}")
            if conn:
                conn.rollback()
            return False
        finally:
            if conn:
                conn.close()

    @staticmethod
    def get_processed_hashes() -> Set[str]:
        conn = None
        try:
            conn = psycopg2.connect(**CONFIG['database'])
            cursor = conn.cursor()
            cursor.execute("SELECT hash FROM log_hashes")
            return {row[0] for row in cursor.fetchall()}
        except Exception as e:
            log_operation(f"Ошибка получения хешей: {e}")
            return set()
        finally:
            if conn:
                conn.close()

    @staticmethod
    @timeit
    def bulk_load(logs: List[dict], progress_callback=None) -> bool:
        if not logs:
            log_operation("Нет данных для загрузки")
            return False
            
        total_logs = len(logs)
        log_operation(f"Начало загрузки {total_logs} записей в БД")
        
        for chunk_start in range(0, total_logs, CONFIG['max_batch_rows']):
            chunk_end = min(chunk_start + CONFIG['max_batch_rows'], total_logs)
            log_chunk = logs[chunk_start:chunk_end]
            
            log_operation(f"Обработка пакета {chunk_start}-{chunk_end} из {total_logs}")
            
            if not DatabaseManager._process_batch(log_chunk, progress_callback, chunk_start, total_logs):
                return False
                
        return True

    @staticmethod
    def _process_batch(logs: List[dict], progress_callback, offset: int, total: int) -> bool:
        conn = None
        cursor = None
        try:
            conn = DatabaseManager._get_connection()
            conn.autocommit = False
            cursor = conn.cursor()
            
            cursor.execute("SET work_mem = '256MB'")
            cursor.execute("SET maintenance_work_mem = '1GB'")
            cursor.execute("SET synchronous_commit TO OFF")
            
            ips = {log['ip'] for log in logs}
            
            cursor.execute("CREATE TEMP TABLE temp_ips (ip VARCHAR(15)) ON COMMIT DROP")
            execute_batch(
                cursor,
                "INSERT INTO temp_ips VALUES (%s)",
                [(ip,) for ip in ips],
                page_size=10000
            )
            
            cursor.execute("""
                INSERT INTO ip_addresses (ip_address, last_seen)
                SELECT ip, NOW() FROM temp_ips
                ON CONFLICT (ip_address) DO UPDATE SET last_seen = EXCLUDED.last_seen
                RETURNING ip_address, id
            """)
            ip_map = {ip: id for ip, id in cursor.fetchall()}
            
            execute_batch(
                cursor,
                "INSERT INTO log_hashes (hash) VALUES (%s) ON CONFLICT DO NOTHING",
                [(log['hash'],) for log in logs],
                page_size=10000
            )
            
            output = io.StringIO()
            for log in logs:
                output.write("\t".join([
                    str(ip_map[log['ip']]),
                    log['datetime'].isoformat(),
                    log['method'],
                    log['api'],
                    '1.1',
                    str(log['status']),
                    str(log['bytes']),
                    '', '', '0',
                    't' if log['is_error'] else 'f',
                    'f'
                ]) + "\n")
            output.seek(0)
            
            cursor.copy_from(
                output,
                'requests',
                columns=('ip_id', 'datetime', 'request_method', 'api_path', 
                        'protocol', 'status_code', 'bytes_sent',
                        'referer', 'user_agent', 'response_time', 'is_error', 'is_attack')
            )
            
            cursor.execute("""
                UPDATE ip_addresses 
                SET 
                    total_requests = subquery.count,
                    is_suspicious = subquery.attack_count > 0
                FROM (
                    SELECT 
                        ip_id, 
                        COUNT(*) as count,
                        SUM(CASE WHEN is_attack THEN 1 ELSE 0 END) as attack_count
                    FROM requests 
                    GROUP BY ip_id
                ) AS subquery
                WHERE ip_addresses.id = subquery.ip_id
            """)
            
            cursor.execute("REFRESH MATERIALIZED VIEW daily_stats")
            conn.commit()
            
            log_operation(f"Успешно загружено {len(logs)} новых записей в пакете")
            return True
            
        except Exception as e:
            log_operation(f"Критическая ошибка при загрузке данных: {e}")
            if conn:
                try:
                    conn.rollback()
                except:
                    pass
            return False
        finally:
            if cursor:
                cursor.close()
            if conn:
                DatabaseManager._return_connection(conn)

    @staticmethod
    @timeit
    def export_to_csv(query: str, filename: str) -> bool:
        conn = None
        cursor = None
        try:
            conn = psycopg2.connect(**CONFIG['database'])
            cursor = conn.cursor()
            cursor.execute(query)
            
            columns = [desc[0] for desc in cursor.description]
            
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                writer.writerow(columns)
                
                while True:
                    rows = cursor.fetchmany(10000)
                    if not rows:
                        break
                    writer.writerows(rows)
            
            return True
        except Exception as e:
            log_operation(f"Ошибка экспорта в CSV: {e}")
            return False
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

    @staticmethod
    @timeit
    def export_to_parquet(query: str, filename: str) -> bool:
        conn = None
        try:
            conn = psycopg2.connect(**CONFIG['database'])
            df = pd.read_sql(query, conn)
            
            for col in df.select_dtypes(include=['datetime64']).columns:
                df[col] = df[col].astype(str)
            
            for col in df.select_dtypes(include=['bool']).columns:
                df[col] = df[col].map({True: 'Yes', False: 'No'})
            
            df.to_parquet(filename, index=False)
            return True
        except Exception as e:
            log_operation(f"Ошибка экспорта в Parquet: {e}")
            return False
        finally:
            if conn:
                conn.close()

    @staticmethod
    def get_db_size() -> str:
        conn = None
        cursor = None
        try:
            conn = psycopg2.connect(**CONFIG['database'])
            cursor = conn.cursor()
            cursor.execute("SELECT pg_size_pretty(pg_database_size(current_database()))")
            return cursor.fetchone()[0]
        except Exception as e:
            log_operation(f"Ошибка получения размера БД: {e}")
            return "N/A"
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

# ==================== МНОГОПОТОЧНАЯ ОБРАБОТКА ====================
class LogProcessor:
    def __init__(self):
        self.seen_hashes = DatabaseManager.get_processed_hashes()
        self.salt = CONFIG['hash_salt']
        self.lock = None

    def process_file(self, file_path: str) -> Tuple[List[dict], int]:
        try:
            logs, new_hashes = FastLogParser.parse_file(
                file_path, 
                self.salt,
                self.seen_hashes
            )
            
            if self.lock:
                with self.lock:
                    self.seen_hashes.update(new_hashes)
            else:
                self.seen_hashes.update(new_hashes)
            
            return logs, len(logs)
        except Exception as e:
            log_operation(f"Ошибка обработки файла {file_path}: {e}")
            return [], 0

class LogProcessorThread(QThread):
    progress = Signal(int, int)
    status = Signal(str)
    finished = Signal(bool)
    db_updated = Signal()
    operation_time = Signal(str)

    def __init__(self, file_paths: List[str]):
        super().__init__()
        self.file_paths = file_paths
        self.start_time = time.time()
        self.processor = LogProcessor()
        self.total_logs = 0
        self.success_files = 0

    def run(self):
        success = True
        total_files = len(self.file_paths)
        all_logs = []
        
        self.processor.seen_hashes = DatabaseManager.get_processed_hashes()
        
        for i, file_path in enumerate(self.file_paths[:3]):
            try:
                with open(file_path, 'rb') as f:
                    raw = f.read(10000)
                    encoding = chardet.detect(raw)['encoding'] or 'utf-8'
                with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                    sample_lines = [next(f).strip() for _ in range(3)]
                log_operation(f"Пример строк из {file_path}:\n" + "\n".join(sample_lines))
            except Exception as e:
                log_operation(f"Не удалось прочитать пример строк из {file_path}: {e}")
        
        total_size = sum(os.path.getsize(f) for f in self.file_paths)
        log_operation(f"Общий размер файлов: {total_size/1024/1024:.2f} MB")
        
        with ThreadPoolExecutor(max_workers=CONFIG['max_workers']) as executor:
            futures = {executor.submit(self.processor.process_file, fp): fp for fp in self.file_paths}
            
            for i, future in enumerate(as_completed(futures), 1):
                file_path = futures[future]
                try:
                    logs, count = future.result()
                    if logs:
                        all_logs.extend(logs)
                        self.total_logs += count
                        self.success_files += 1
                    
                    self.status.emit(
                        f"Обработано {i}/{total_files} файлов. Найдено {count} новых записей"
                    )
                    self.progress.emit(i, total_files)
                except Exception as e:
                    self.status.emit(f"Ошибка обработки файла {file_path}: {str(e)}")
                    success = False
        
        if all_logs:
            self.status.emit(f"Загрузка {len(all_logs)} записей в БД...")
            QApplication.processEvents()
            
            load_success = DatabaseManager.bulk_load(
                all_logs,
                lambda c, t: self.progress.emit(c, t)
            )
            
            success = success and load_success
        
        operation_time = time.time() - self.start_time
        if success:
            self.db_updated.emit()
            self.operation_time.emit(
                f"Обработано {self.success_files}/{total_files} файлов, {self.total_logs} записей за {operation_time:.2f} сек"
            )
        else:
            self.operation_time.emit(
                f"Обработано {self.success_files}/{total_files} файлов с ошибками, {self.total_logs} записей за {operation_time:.2f} сек"
            )
        self.finished.emit(success)

# ==================== ГРАФИЧЕСКИЙ ИНТЕРФЕЙС ====================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Анализатор логов веб-сервера Дети-GPT")
        self.setup_ui()
        self.setup_db()
        self.resize(1200, 800)

    def export_ip_table_to_excel(self):
        try:
            now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = os.path.join(os.getcwd(), f"ip_stats_{now}.xlsx")
            conn = psycopg2.connect(**CONFIG['database'])
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT DISTINCT DATE(datetime) as day 
                FROM requests 
                ORDER BY day
            """)
            days = [row[0] for row in cursor.fetchall()]
            
            cursor.execute("""
                SELECT 
                    ip.ip_address,
                    DATE(r.datetime) as day,
                    COUNT(*) as requests
                FROM requests r
                JOIN ip_addresses ip ON r.ip_id = ip.id
                GROUP BY ip.ip_address, DATE(r.datetime)
                ORDER BY requests DESC
                LIMIT 1000
            """)
            data = cursor.fetchall()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "IP Statistics"
            
            headers = ["IP Address"] + [str(day) for day in days]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}1"] = header
                ws[f"{col_letter}1"].font = Font(bold=True)
                ws[f"{col_letter}1"].alignment = Alignment(horizontal='center')
            ip_data = {}
            for ip, day, count in data:
                if ip not in ip_data:
                    ip_data[ip] = {}
                ip_data[ip][day] = count
        
            for row_num, ip in enumerate(ip_data.keys(), 2):
                ws[f"A{row_num}"] = ip
                for col_num, day in enumerate(days, 2):
                    col_letter = get_column_letter(col_num)
                    count = ip_data[ip].get(day, 0)
                    ws[f"{col_letter}{row_num}"] = count
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filename)
            
            QMessageBox.information(
                self, 
                "Экспорт завершен", 
                f"Таблица IP по дням успешно экспортирована в файл:\n{filename}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Ошибка", 
                f"Не удалось экспортировать таблицу IP:\n{str(e)}"
            )
        finally:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()

    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        load_group = QWidget()
        load_layout = QVBoxLayout(load_group)
        
        self.btn_load_file = QPushButton("Загрузить файл лога")
        self.btn_load_file.clicked.connect(self.load_file)
        load_layout.addWidget(self.btn_load_file)

        self.btn_load_folder = QPushButton("Загрузить папку с логами")
        self.btn_load_folder.clicked.connect(self.load_folder)
        load_layout.addWidget(self.btn_load_folder)

        layout.addWidget(load_group)

        export_group = QWidget()
        export_layout = QVBoxLayout(export_group)
        
        self.btn_export_csv = QPushButton("Экспорт в CSV")
        self.btn_export_csv.clicked.connect(lambda: self.export_data('csv'))
        export_layout.addWidget(self.btn_export_csv)

        self.btn_export_parquet = QPushButton("Экспорт в Parquet")
        self.btn_export_parquet.clicked.connect(lambda: self.export_data('parquet'))
        export_layout.addWidget(self.btn_export_parquet)

        layout.addWidget(export_group)

        viz_group = QWidget()
        viz_layout = QVBoxLayout(viz_group)
        
        self.btn_plot_requests = QPushButton("График запросов")
        self.btn_plot_requests.clicked.connect(self.plot_requests)
        viz_layout.addWidget(self.btn_plot_requests)

        self.btn_plot_errors = QPushButton("График ошибок")
        self.btn_plot_errors.clicked.connect(self.plot_errors)
        viz_layout.addWidget(self.btn_plot_errors)

        self.btn_ip_table = QPushButton("IP по дням (таблица)")
        self.btn_ip_table.clicked.connect(self.show_ip_table)
        viz_layout.addWidget(self.btn_ip_table)

        self.btn_export_ip_excel = QPushButton("Экспорт IP по дням (Excel)")
        self.btn_export_ip_excel.clicked.connect(self.export_ip_table_to_excel)
        viz_layout.addWidget(self.btn_export_ip_excel)

        layout.addWidget(viz_group)

        info_group = QWidget()
        info_layout = QVBoxLayout(info_group)
        
        self.db_size_label = QLabel(f"Размер БД: {DatabaseManager.get_db_size()}")
        info_layout.addWidget(self.db_size_label)

        self.time_label = QLabel("Время операций: -")
        info_layout.addWidget(self.time_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        info_layout.addWidget(self.progress_bar)

        self.status_label = QLabel("Готов к работе")
        self.status_label.setAlignment(Qt.AlignCenter)
        info_layout.addWidget(self.status_label)

        layout.addWidget(info_group)

        self.table = QTableWidget()
        self.table.setColumnCount(0)
        self.table.setRowCount(0)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table)

    def setup_db(self):
        try:
            if DatabaseManager.clear_database() and DatabaseManager.create_schema():
                self.update_db_info()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось инициализировать БД:\n{str(e)}")

    def load_file(self):
        file, _ = QFileDialog.getOpenFileName(
            self, 
            "Выберите файл лога", 
            "", 
            "Логи (*.log *.txt);;Все файлы (*)"
        )
        if file:
            self.process_files([file])

    def load_folder(self):
        folder = QFileDialog.getExistingDirectory(
            self, 
            "Выберите папку с логами"
        )
        if folder:
            files = (
                glob.glob(os.path.join(folder, "*.log")) + 
                glob.glob(os.path.join(folder, "*.txt"))
            )
            if files:
                self.process_files(files)

    def process_files(self, files: List[str]):
        if not files:
            QMessageBox.warning(self, "Ошибка", "Не выбраны файлы для обработки")
            return
            
        log_operation(f"Начата обработка {len(files)} файлов")
        
        sample_file = files[0]
        try:
            with open(sample_file, 'rb') as f:
                raw = f.read(10000)
                encoding = chardet.detect(raw)['encoding'] or 'utf-8'
            with open(sample_file, 'r', encoding=encoding, errors='ignore') as f:
                sample_lines = [next(f).strip() for _ in range(5)]
            log_operation(f"Пример строк из {sample_file}:\n" + "\n".join(sample_lines))
        except Exception as e:
            log_operation(f"Не удалось прочитать пример строк: {e}")
        
        self.thread = LogProcessorThread(files)
        self.thread.progress.connect(self.update_progress)
        self.thread.status.connect(self.update_status)
        self.thread.finished.connect(self.on_processing_finished)
        self.thread.db_updated.connect(self.update_db_info)
        self.thread.operation_time.connect(self.update_time_label)
        
        self.set_buttons_enabled(False)
        self.status_label.setText("Начата обработка логов...")
        self.thread.start()

    def update_progress(self, current: int, total: int):
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)

    def update_status(self, message: str):
        self.status_label.setText(message)

    def update_time_label(self, message: str):
        self.time_label.setText(message)

    def update_db_info(self):
        self.db_size_label.setText(f"Размер БД: {DatabaseManager.get_db_size()}")

    def on_processing_finished(self, success: bool):
        self.set_buttons_enabled(True)
        if success:
            QMessageBox.information(self, "Успех", "Обработка логов завершена!")
            self.status_label.setText("Готов к работе")
        else:
            QMessageBox.warning(self, "Ошибка", "Обработка завершена с ошибками")
            self.status_label.setText("Ошибка обработки")

    def export_data(self, format: str):
        file, _ = QFileDialog.getSaveFileName(
            self,
            f"Экспорт в {format.upper()}",
            "",
            f"{format.upper()} файлы (*.{format})"
        )
        if not file:
            return
        
        query = """
            SELECT 
                ip.ip_address AS "IP Address",
                req.datetime AS "Date Time",
                req.request_method AS "Method",
                req.api_path AS "API Path",
                req.protocol AS "Protocol",
                req.status_code AS "Status Code",
                req.bytes_sent AS "Bytes Sent",
                req.referer AS "Referer",
                req.user_agent AS "User Agent",
                req.response_time AS "Response Time (ms)",
                CASE WHEN req.is_error THEN 'Yes' ELSE 'No' END AS "Is Error",
                CASE WHEN req.is_attack THEN 'Yes' ELSE 'No' END AS "Is Attack"
            FROM requests req
            JOIN ip_addresses ip ON req.ip_id = ip.id
            ORDER BY req.datetime
        """
        
        try:
            start_time = time.time()
            self.status_label.setText(f"Экспорт в {format.upper()}...")
            QApplication.processEvents()
            
            if format == 'csv':
                success = DatabaseManager.export_to_csv(query, file)
            elif format == 'parquet':
                success = DatabaseManager.export_to_parquet(query, file)
            else:
                QMessageBox.warning(self, "Ошибка", "Неподдерживаемый формат экспорта")
                return
            
            operation_time = time.time() - start_time
            if success:
                self.time_label.setText(f"Экспорт завершен за {operation_time:.2f} сек")
                QMessageBox.information(
                    self,
                    "Успех",
                    f"Данные успешно экспортированы в {format.upper()}\n"
                    f"Время выполнения: {operation_time:.2f} сек\n"
                    f"Размер файла: {os.path.getsize(file)/1024/1024:.2f} MB"
                )
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось экспортировать данные")
            
            self.status_label.setText("Готов к работе")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта: {str(e)}")
            self.status_label.setText("Ошибка экспорта")

    def plot_requests(self):
        try:
            start_time = time.time()
            self.status_label.setText("Построение графика запросов...")
            QApplication.processEvents()
            
            conn = psycopg2.connect(**CONFIG['database'])
            df = pd.read_sql("SELECT day, total_requests FROM daily_stats", conn)
            conn.close()
            
            plt.figure(figsize=(12, 6))
            plt.bar(df['day'], df['total_requests'], color='skyblue')
            plt.title("Количество запросов по дням", fontsize=14)
            plt.xlabel("Дата", fontsize=12)
            plt.ylabel("Количество запросов", fontsize=12)
            plt.grid(axis='y', linestyle='--', alpha=0.7)
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.show()
            
            operation_time = time.time() - start_time
            self.time_label.setText(f"График построен за {operation_time:.2f} сек")
            self.status_label.setText("Готов к работе")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось построить график:\n{str(e)}")
            self.status_label.setText("Ошибка построения графика")

    def plot_errors(self):
        try:
            start_time = time.time()
            self.status_label.setText("Построение графика ошибок...")
            QApplication.processEvents()
            
            conn = psycopg2.connect(**CONFIG['database'])
            df = pd.read_sql("""
                SELECT 
                    status_code,
                    COUNT(*) as count,
                    CASE 
                        WHEN status_code BETWEEN 400 AND 499 THEN '4xx'
                        WHEN status_code >= 500 THEN '5xx'
                        ELSE 'Другие'
                    END as error_group
                FROM requests
                WHERE status_code >= 400
                GROUP BY status_code, error_group
                ORDER BY count DESC
            """, conn)
            conn.close()
            
            plt.figure(figsize=(12, 6))
            
            plt.subplot(1, 2, 1)
            plt.bar(df['status_code'].astype(str), df['count'], color='salmon')
            plt.title("Распределение по кодам", fontsize=12)
            plt.xlabel("Код статуса", fontsize=10)
            plt.ylabel("Количество", fontsize=10)
            
            plt.subplot(1, 2, 2)
            df_grouped = df.groupby('error_group').sum().reset_index()
            plt.pie(
                df_grouped['count'], 
                labels=df_grouped['error_group'], 
                autopct='%1.1f%%',
                colors=['lightcoral', 'indianred', 'firebrick']
            )
            plt.title("Распределение по группам", fontsize=12)
            
            plt.suptitle("Статистика ошибок сервера", fontsize=14)
            plt.tight_layout()
            plt.show()
            
            operation_time = time.time() - start_time
            self.time_label.setText(f"График построен за {operation_time:.2f} сек")
            self.status_label.setText("Готов к работе")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось построить график:\n{str(e)}")
            self.status_label.setText("Ошибка построения графика")

    def show_ip_table(self):
        self.table.clear()
        self.status_label.setText("Подготовка таблицы...")
        QApplication.processEvents()
        
        QTimer.singleShot(100, self._load_ip_table)

    def _load_ip_table(self):
        conn = None
        cursor = None
        try:
            start_time = time.time()
            self.status_label.setText("Загрузка данных...")
            QApplication.processEvents()
            
            conn = psycopg2.connect(**CONFIG['database'])
            cursor = conn.cursor()
            
            cursor.execute(f"""
                SELECT 
                    ip.ip_address, 
                    DATE(r.datetime) as day, 
                    COUNT(*) as requests
                FROM requests r
                JOIN ip_addresses ip ON r.ip_id = ip.id
                GROUP BY ip.ip_address, DATE(r.datetime)
                ORDER BY requests DESC
                LIMIT {CONFIG['max_table_rows']}
            """)
            
            data = cursor.fetchall()
            if not data:
                self.status_label.setText("Нет данных для отображения")
                QMessageBox.information(self, "Информация", "Нет данных для таблицы")
                return
            
            ips = sorted({row[0] for row in data})
            days = sorted({row[1] for row in data})
            
            self.table.setRowCount(len(ips))
            self.table.setColumnCount(len(days) + 1)
            self.table.setHorizontalHeaderLabels(['IP'] + [str(day) for day in days])
            
            ip_day_counts = {(row[0], row[1]): row[2] for row in data}
            for i, ip in enumerate(ips):
                self.table.setItem(i, 0, QTableWidgetItem(ip))
                for j, day in enumerate(days, 1):
                    count = ip_day_counts.get((ip, day), 0)
                    item = QTableWidgetItem(str(count))
                    
                    if count > 1000:
                        item.setBackground(Qt.yellow)
                    
                    self.table.setItem(i, j, item)
            
            self.table.resizeColumnsToContents()
            
            operation_time = time.time() - start_time
            self.time_label.setText(f"Таблица загружена за {operation_time:.2f} сек")
            self.status_label.setText("Готов к работе")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить таблицу:\n{str(e)}")
            self.status_label.setText("Ошибка загрузки таблицы")
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

    def set_buttons_enabled(self, enabled: bool):
        for btn in [
            self.btn_load_file, self.btn_load_folder,
            self.btn_export_csv, self.btn_export_parquet,
            self.btn_plot_requests, self.btn_plot_errors,
            self.btn_ip_table
        ]:
            btn.setEnabled(enabled)

if __name__ == "__main__":
    app = QApplication([])
    app.setStyle('Fusion')
    
    window = MainWindow()
    window.show()
    
    app.exec_()
